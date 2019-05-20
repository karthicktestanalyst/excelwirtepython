from selenium import webdriver
from selenium.common.exceptions import WebDriverException
#import xlsxwriter
from openpyxl import Workbook
import openpyxl
import time
from selenium.webdriver.common.action_chains import ActionChains
from time import sleep
from pip._internal.utils.ui import hidden_cursor



driver = webdriver.Chrome('E:\Selenium\Driver\chromedriver.exe') 
driver.maximize_window()
driver.get('https://www.flipkart.com/')
driver.find_element_by_xpath('//button[@class="_2AkmmA _29YdH8"]').click()
driver.find_element_by_xpath('//span[text()="TVs & Appliances"]').click()
driver.implicitly_wait(30)
driver.find_element_by_xpath('//a[text()="Electric Cookers"]').click()

act = ActionChains(driver)
parGUID = driver.current_window_handle
wb = openpyxl.Workbook() 
sheet = wb.active
nrow=1

  

#len(tottalElt)
for p in range(0,10):
    tottalElt = driver.find_elements_by_xpath('//a[@class="_2cLu-l"]')
    for c in range(0, len(tottalElt)):
        staticPart = "//a[@class='_2cLu-l'][text()='"
        forPass = staticPart + tottalElt[c].text + "']"
    
        try:
            driver.find_element_by_xpath(forPass).click()
            for guid in driver.window_handles:
                if guid != parGUID:
                    driver.switch_to_window(guid)
                    name = driver.find_element_by_xpath('//span[@class="_35KyD6"]').text
                    price = driver.find_element_by_xpath('//div[@class="_1vC4OE _3qQ9m1"]').text
                    print name
                    print c
                    sheet.cell(row=nrow, column=1).value = name
                    sheet.cell(row=nrow, column=2).value = price
                    nrow+=1
                    sleep(2)
                    driver.close()
                    driver.switch_to_window(parGUID)
            
        except WebDriverException as e:
            print "WebDriverException"

    driver.find_element_by_xpath('//a[@class="_3fVaIS"]').click()
wb.save("D:\\sample.xls") 
