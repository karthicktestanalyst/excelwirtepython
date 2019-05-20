from selenium import webdriver
from cgitb import text

import openpyxl 
from argparse import Action
from selenium.webdriver.common.action_chains import ActionChains
import string
from selenium.common.exceptions import WebDriverException
from time import sleep

driver = webdriver.Chrome('E:\Selenium\Driver\chromedriver.exe') 
driver.maximize_window()
driver.get('https://www.flipkart.com/')
driver.find_element_by_xpath('//button[@class="_2AkmmA _29YdH8"]').click()
driver.find_element_by_xpath('//span[text()="TVs & Appliances"]').click()
driver.implicitly_wait(30)
driver.find_element_by_xpath('//a[text()="Refrigerators"]').click()
driver.find_element_by_xpath('//a[@class="_3XS1AH _32ZSYo"]').click()

action = ActionChains(driver)

parGUID = driver.current_window_handle
wb = openpyxl.Workbook() 
sheet = wb.active
rowNum = 1
# loop for page navigation
for p in range(0,2):
    #loop for details fetching and stored in variable
    tottalelt = driver.find_elements_by_xpath('//div[@class="_3wU53n"]')
    for i in range(0, len(tottalelt)) :
        count = "//div[@class='_3wU53n'][text()='"
        fridge = count + tottalelt[i].text + "']"
        
        try:
            driver.find_element_by_xpath(fridge).click()
            for guid in driver.window_handles:
                if guid != parGUID:
                    driver.switch_to_window(guid)
                    namee=driver.find_element_by_xpath('//span[@class="_35KyD6"]').text
                    price=driver.find_element_by_xpath('//div[@class="_1vC4OE _3qQ9m1"]').text
                    rating = driver.find_element_by_xpath('//div[@class="hGSR34"]').text
                    highlights=driver.find_element_by_xpath('//div[@class="_3WHvuP"]').text
                    sheet.cell(row=rowNum, column=1).value = namee
                    sheet.cell(row=rowNum, column=2).value = price
                    sheet.cell(row=rowNum, column=3).value = rating
                    sheet.cell(row=rowNum, column=4).value =highlights
                    rowNum += 1
                    sleep(2)
                    driver.close()
                driver.switch_to_window(parGUID)
                
                
        except WebDriverException as e:
            print "WebDriverException"
            # navigate to next page by clicking next button
    driver.find_element_by_xpath("//span[text()='Next']").click()
    sleep(3)
    #fetching values stored in excel sheet in respective columns
wb.save("D:\\refrigerators.xlsx")
print "FINISHED"
driver.close()

 


